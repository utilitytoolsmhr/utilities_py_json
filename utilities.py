import json
import os
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def get_json_files(directory):
    return [f for f in os.listdir(directory) if f.endswith('.json')]

def get_defaults_structure(json_data):
    def process_dict(data):
        if isinstance(data, dict):
            return {k: process_dict(v) for k, v in data.items()}
        elif isinstance(data, list):
            return [process_dict(data[0])] if data else []
        elif isinstance(data, str):
            return ""
        elif isinstance(data, int):
            return 0
        elif isinstance(data, float):
            return 0.0
        else:
            return None
    return process_dict(json_data)

def write_to_excel(headers, sheet):
    sheet["A1"] = "Campo"
    col = 1
    for header in headers:
        sheet[f"{get_column_letter(col)}2"] = header
        col += 1

def write_defaults_to_json(defaults, output_path):
    with open(output_path, 'w') as f:
        json.dump(defaults, f, indent=4)

def format_value(value):
    if isinstance(value, str):
        return '""'
    elif isinstance(value, int):
        return '0'
    elif isinstance(value, float):
        return '0.0'
    return 'None'

def format_json(json_obj, path='', lines=None):
    if lines is None:
        lines = []
    if isinstance(json_obj, dict):
        for k, v in json_obj.items():
            new_path = f"{path}['{k}']" if path else f"['{k}']"
            format_json(v, new_path, lines)
    elif isinstance(json_obj, list):
        if len(json_obj) > 0:
            new_path = f"{path}[0]"
            format_json(json_obj[0], new_path, lines)
    else:
        lines.append(f"    json_return{path} = None")
    return lines

def create_headers(sheet, max_col):
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col + 6)
    sheet.cell(row=2, column=max_col + 1, value="Tipo").fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    sheet.cell(row=2, column=max_col + 2, value="Valores").fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    sheet.merge_cells(start_row=2, start_column=max_col + 3, end_row=2, end_column=max_col + 7)
    sheet.cell(row=2, column=max_col + 3, value="Mapeo").fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    sheet.cell(row=3, column=max_col + 3, value="DOMINIO DF")
    sheet.cell(row=3, column=max_col + 4, value="SUBDOMINIO DF")
    sheet.cell(row=3, column=max_col + 5, value="PROPOSITO")
    sheet.cell(row=3, column=max_col + 6, value="CAMPO")
    sheet.cell(row=3, column=max_col + 7, value="ATRIBUTO")
    sheet.cell(row=2, column=max_col + 8, value="Agregado").fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    sheet.cell(row=2, column=max_col + 9, value="Error 403").fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    sheet.cell(row=2, column=max_col + 10, value="Observaciones").fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")

def adjust_column_width(sheet):
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

def get_max_column_width(json_data, col=1):
    max_col = col
    if isinstance(json_data, dict):
        for key, value in json_data.items():
            if isinstance(value, (dict, list)):
                new_col = get_max_column_width(value, col + 1)
                max_col = max(max_col, new_col)
    elif isinstance(json_data, list) and len(json_data) > 0 and isinstance(json_data[0], (dict, list)):
        new_col = get_max_column_width(json_data[0], col + 1)
        max_col = max(max_col, new_col)
    return max_col

def write_data_to_excel(json_data, sheet):
    current_row = 4
    max_col = get_max_column_width(json_data)

    def process_dict(data, col, row):
        for key, value in data.items():
            sheet[f"{get_column_letter(col)}{row}"] = key
            if isinstance(value, dict):
                row = process_dict(value, col + 1, row + 1)
            elif isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
                row = process_dict(value[0], col + 1, row + 1)
            else:
                if isinstance(value, int):
                    data_type = "Integer"
                elif isinstance(value, float):
                    data_type = "Float"
                elif isinstance(value, str):
                    data_type = "String"
                elif isinstance(value, list):
                    data_type = "List"
                    value = ', '.join(map(str, value))
                else:
                    data_type = "Unknown"
                    value = str(value)
                sheet[f"{get_column_letter(max_col + 1)}{row}"] = data_type
                sheet[f"{get_column_letter(max_col + 2)}{row}"] = value
                row += 1
        return row
    
    process_dict(json_data, 1, current_row)
    create_headers(sheet, max_col)
    adjust_column_width(sheet)

def process_json_file(json_path):
    with open(json_path, 'r') as file:
        data = json.load(file)
    base_name = 'json_return_' + os.path.splitext(os.path.basename(json_path))[0]
    cr_json_name = f"cr_{os.path.basename(json_path)}"
    cr_json_path = os.path.join(os.path.dirname(json_path), cr_json_name)
    if not os.path.exists(cr_json_path):
        defaults = get_defaults_structure(data)
        write_defaults_to_json(defaults, cr_json_path)
    with open(cr_json_path, 'r') as file:
        cr_data = json.load(file)
    lines = []
    formatted_json = json.dumps(data)
    lines.append(f"{base_name} = {formatted_json}")
    lines.append(f"json_return = json.loads({base_name})")
    lines.append("")
    lines.append(f"cr_{base_name} = {json.dumps(cr_data)}")
    lines.append("")
    lines.append("def tfn(fna, fti):")
    lines.append("    pass")
    lines.append("")
    lines.append("def main(p):")
    lines.append("")
    replacement_lines = format_json(data)
    lines.extend(replacement_lines)
    lines.append("    return json_return")
    lines.append("")
    new_file_path = f"{os.path.splitext(os.path.basename(json_path))[0]}.py"
    with open(new_file_path, 'w') as new_file:
        new_file.write('\n'.join(lines))

def main(P):
    workbook = openpyxl.Workbook()
    for filename in os.listdir(P):
        if filename.endswith('.json'):
            json_path = os.path.join(P, filename)
            with open(json_path, 'r') as f:
                json_data = json.load(f)
                sheet_name = os.path.splitext(filename)[0][:31]
                sheet = workbook.create_sheet(title=sheet_name)
                write_data_to_excel(json_data, sheet)
                process_json_file(json_path)
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    workbook.save(os.path.join(P, "headers.xlsx"))

if __name__ == "__main__":
    main(".")
